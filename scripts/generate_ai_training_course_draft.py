from __future__ import annotations

import csv
import json
import re
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path("/Users/haichao/Desktop/work/51ToolBox")
SOURCE_XLSX = Path("/Users/haichao/Downloads/AI培训课程需求.xlsx")
OUTPUT_DIR = ROOT / "deliverables" / "ai_training_course_draft"
FRAGMENTS_DIR = OUTPUT_DIR / "fragments"


REQUIRED_OPTION_FIELDS = (
    "option_type",
    "course_name",
    "course_outline",
    "course_intro",
    "teacher_name",
    "teacher_bio",
)
EXPECTED_OPTION_TYPES = ("A", "B", "C", "D")
OPTION_TITLES = {
    "A": "方法论版",
    "B": "实战版",
    "C": "专项深化版",
    "D": "方案共创版",
}
OPTION_DELIVERY_FOCUS = {
    "A": "关键方法、流程框架与质量控制点验证",
    "B": "现场搭建、联调调试与结果验证",
    "C": "专题能力深化、参数优化与专项实验",
    "D": "分组共创、方案产出与专家答辩",
}
OPTION_HANDS_ON_SECTION_TITLE = {
    "A": "实训安排：方法验证与关键链路演示",
    "B": "实操安排：现场上机联调与结果验证",
    "C": "专项实验：专题能力深化与性能观察",
    "D": "共创演练：分组上机实训与方案答辩",
}

PROJECT_REQUIREMENTS = {
    "AI应用开发": {
        "session_schedule": "14期，每期3天",
        "participant_range": "每场约30-60人",
        "delivery_mode": "线下培训模式",
        "training_audience": "研发、测试人员等",
        "training_objective": "掌握AI应用低代码与标准开发流程，提升模型集成与系统对接技术能力，建立AI应用质量测评与安全保障意识。",
        "requires_on_site_teaching": True,
        "requires_hands_on": True,
        "requires_environment": True,
        "delivery_requirement": "培训采用现场授课与上机实操演练相结合的方式开展，需在我司培训场地布设或提供可供电脑上机实操的环境。",
        "teacher_standard": "讲师应具备头部科技公司或同等技术能力公司的AI平台或应用建设背景，并拥有AI领域合作成功落地经验。",
        "practice_focus": "AI应用低代码开发、模型集成、系统对接、质量测评与安全验证",
        "teacher_delivery_suffix": "同时具备现场授课、上机实操组织与工程辅导经验，可在甲方培训场地完成从讲解、演示到上机验证的完整交付。",
    },
    "AI应用解决方案": {
        "session_schedule": "7期，每期1天",
        "participant_range": "每场约30-60人",
        "delivery_mode": "线下培训模式",
        "training_audience": "产品经理、解决方案及需求分析人员等",
        "training_objective": "掌握AI场景识别与业务需求分析方法，具备AI解决方案架构设计与业务建模能力，提升跨部门协作与方案落地推动能力。",
        "requires_on_site_teaching": False,
        "requires_hands_on": False,
        "requires_environment": False,
        "delivery_requirement": "培训以线下授课为主，可结合案例研讨、情景演练与方案演示推进课堂互动。",
        "teacher_standard": "讲师应具备AI产品规划、解决方案设计或金融场景建模经验，能够支撑跨部门协作与方案落地辅导。",
        "practice_focus": "场景识别、需求转化、业务建模与方案论证",
        "teacher_delivery_suffix": "",
    },
    "AI应用运营": {
        "session_schedule": "7期，每期1天",
        "participant_range": "每场约30-60人",
        "delivery_mode": "线下培训模式",
        "training_audience": "运营、数据分析人员等",
        "training_objective": "掌握AI应用智能运维与性能监控方法，提升运营数据分析与价值验证能力，建立AI应用风险识别与控制机制。",
        "requires_on_site_teaching": False,
        "requires_hands_on": False,
        "requires_environment": False,
        "delivery_requirement": "培训以线下授课为主，可结合案例推演、情景演练和操作示范帮助学员理解运营方法。",
        "teacher_standard": "讲师应具备AI运营、数据分析或金融应用运营实践经验，能够结合运营场景讲授指标分析与风险控制。",
        "practice_focus": "智能运维、性能监控、数据分析与风险控制",
        "teacher_delivery_suffix": "",
    },
    "AI模型算法工程化": {
        "session_schedule": "1期，每期3天",
        "participant_range": "每场约30-60人",
        "delivery_mode": "线下培训模式",
        "training_audience": "算法工程师、机器学习工程师等",
        "training_objective": "掌握模型架构设计与算法选择原则，提升模型训练调优与性能优化技能，熟悉模型部署上线与生命周期管理。",
        "requires_on_site_teaching": True,
        "requires_hands_on": True,
        "requires_environment": True,
        "delivery_requirement": "培训采用现场授课与上机实操演练相结合的方式开展，需在我司培训场地布设或提供可供电脑上机实操的环境。",
        "teacher_standard": "讲师应具备头部科技公司或同等技术能力公司的模型工程、训练平台或部署优化背景，并拥有AI领域合作成功落地经验。",
        "practice_focus": "模型训练调优、性能优化、部署上线与生命周期管理",
        "teacher_delivery_suffix": "同时具备训练环境演示、上机实操辅导和模型部署验证经验，可指导学员在现场完成训练、评测或部署类实验。",
    },
    "AI数据能力提升": {
        "session_schedule": "9期，每期2天",
        "participant_range": "每场约30-60人",
        "delivery_mode": "线下培训模式",
        "training_audience": "数据管理、数据工程师等",
        "training_objective": "掌握AI数据采集、清洗与标注规范，建立数据治理框架与安全管理体系，提升数据质量保障与价值挖掘能力。",
        "requires_on_site_teaching": True,
        "requires_hands_on": True,
        "requires_environment": True,
        "delivery_requirement": "培训采用现场授课与上机实操演练相结合的方式开展，需在我司培训场地布设或提供可供电脑上机实操的环境。",
        "teacher_standard": "讲师应具备头部科技公司或同等技术能力公司的数据平台、数据工程或AI数据治理背景，并拥有AI领域合作成功落地经验。",
        "practice_focus": "数据采集、清洗、标注、治理、安全管理与价值挖掘",
        "teacher_delivery_suffix": "同时具备现场数据工程演练、工具操作指导与实训环境组织经验，可带领学员完成从数据处理到治理验证的完整上机过程。",
    },
    "AI能力全员赋能": {
        "session_schedule": "7期，每期1天",
        "participant_range": "每场约50-100人",
        "delivery_mode": "线下培训模式",
        "training_audience": "全员",
        "training_objective": "理解AI与生成式AI（AIGC）的基本概念、能力边界与发展现状，了解AI在金融业特别是银行业务与科技领域的主流应用场景与趋势。建立对AI技术影响个人工作与团队协作的初步前瞻性思考，掌握与AI高效对话的核心方法，筑牢在金融工作中应用AI的数据安全、合规与伦理底线意识。",
        "requires_on_site_teaching": False,
        "requires_hands_on": False,
        "requires_environment": False,
        "delivery_requirement": "培训以线下授课为主，可结合场景演示、操作示范和互动答疑提升全员理解与应用认知。",
        "teacher_standard": "讲师应具备AI通识培训与金融场景应用推广经验，能够将技术概念转化为员工易理解、可执行的方法。",
        "practice_focus": "AI通识认知、办公提效、场景理解与合规使用",
        "teacher_delivery_suffix": "",
    },
}


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def normalize_duration(duration_raw: str) -> str:
    text = normalize_text(duration_raw)
    if not text:
        return ""
    if text == "半天":
        return "4小时（半天）"
    if text.endswith("小时"):
        return text
    if text.endswith("天"):
        match = re.match(r"(\d+(?:\.\d+)?)天", text)
        if match:
            days = float(match.group(1))
            hours = int(days * 8) if days.is_integer() else days * 8
            return f"{hours}小时（{text}）"
    if "/" in text and "小时" in text:
        return text
    if re.fullmatch(r"\d+(?:\.\d+)?", text):
        return f"{text}小时"
    return text


def build_project_requirement_summary(requirement: dict[str, Any]) -> str:
    lines = [
        f"培训方式：{requirement['delivery_mode']}",
        f"培训安排：{requirement['session_schedule']}，{requirement['participant_range']}",
        f"培训对象：{requirement['training_audience']}",
        f"培训目标：{requirement['training_objective']}",
        f"实施要求：{requirement['delivery_requirement']}",
        f"师资要求：{requirement['teacher_standard']}",
    ]
    return "；".join(lines)


def enrich_topic(topic: dict[str, Any]) -> dict[str, Any]:
    requirement = PROJECT_REQUIREMENTS[topic["project"]]
    enriched = dict(topic)
    enriched.update(
        {
            "project_session_schedule": requirement["session_schedule"],
            "project_participant_range": requirement["participant_range"],
            "project_delivery_mode": requirement["delivery_mode"],
            "project_training_audience": requirement["training_audience"],
            "project_training_objective": requirement["training_objective"],
            "project_requires_on_site_teaching": requirement["requires_on_site_teaching"],
            "project_requires_hands_on": requirement["requires_hands_on"],
            "project_requires_environment": requirement["requires_environment"],
            "project_delivery_requirement": requirement["delivery_requirement"],
            "project_teacher_standard": requirement["teacher_standard"],
            "project_practice_focus": requirement["practice_focus"],
            "project_requirement_summary": build_project_requirement_summary(requirement),
        }
    )
    return enriched


def extract_topics() -> list[dict[str, Any]]:
    workbook = load_workbook(SOURCE_XLSX, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows: list[dict[str, Any]] = []
    current_project = ""
    current_period = ""
    for row_idx in range(3, worksheet.max_row + 1):
        seq = worksheet.cell(row_idx, 1).value
        if seq is None:
            continue
        project_cell = normalize_text(worksheet.cell(row_idx, 2).value)
        period_cell = normalize_text(worksheet.cell(row_idx, 3).value)
        if project_cell:
            current_project = project_cell
        if period_cell:
            current_period = period_cell
        duration_raw = normalize_text(worksheet.cell(row_idx, 8).value)
        if not duration_raw and int(seq) in (3, 4):
            duration_raw = "半天"
        topic = {
            "seq": int(seq),
            "project": current_project,
            "project_period": current_period,
            "topic": normalize_text(worksheet.cell(row_idx, 4).value),
            "audience": normalize_text(worksheet.cell(row_idx, 5).value),
            "objective": normalize_text(worksheet.cell(row_idx, 6).value),
            "source_outline": normalize_text(worksheet.cell(row_idx, 7).value),
            "duration_raw": duration_raw,
            "duration_normalized": normalize_duration(duration_raw),
            "teacher_requirement": normalize_text(worksheet.cell(row_idx, 9).value),
        }
        rows.append(enrich_topic(topic))
    return rows


def load_fragments() -> tuple[dict[int, dict[str, Any]], int]:
    topic_map: dict[int, dict[str, Any]] = {}
    fragment_paths = sorted(FRAGMENTS_DIR.glob("*.json")) if FRAGMENTS_DIR.exists() else []
    if fragment_paths:
        for fragment_path in fragment_paths:
            with fragment_path.open("r", encoding="utf-8") as handle:
                fragment = json.load(handle)
            topics = fragment if isinstance(fragment, list) else fragment.get("topics", [])
            for topic in topics:
                seq = int(topic["seq"])
                merged_topic = topic_map.setdefault(seq, {"seq": seq, "options": {}})
                for option in topic.get("options", []):
                    option_type = normalize_text(option.get("option_type"))
                    if option_type in merged_topic["options"]:
                        raise ValueError(
                            f"发现重复课程方案: seq={seq} option={option_type} ({fragment_path.name})"
                        )
                    merged_topic["options"][option_type] = option
    else:
        workbook = load_workbook(SOURCE_XLSX, data_only=True)
        worksheet = workbook[workbook.sheetnames[0]]
        current_seq: int | None = None
        for row_idx in range(3, worksheet.max_row + 1):
            seq = worksheet.cell(row_idx, 1).value
            if seq is not None:
                current_seq = int(seq)
            if current_seq is None:
                continue
            course_name = normalize_text(worksheet.cell(row_idx, 10).value)
            option_marker = normalize_text(worksheet.cell(row_idx, 15).value)
            if not course_name or not option_marker:
                continue
            option_type = option_marker[-1]
            merged_topic = topic_map.setdefault(current_seq, {"seq": current_seq, "options": {}})
            merged_topic["options"][option_type] = {
                "option_type": option_type,
                "course_name": course_name,
                "course_outline": normalize_text(worksheet.cell(row_idx, 11).value),
                "course_intro": normalize_text(worksheet.cell(row_idx, 12).value),
                "teacher_name": normalize_text(worksheet.cell(row_idx, 13).value),
                "teacher_bio": normalize_text(worksheet.cell(row_idx, 14).value),
            }
    finalized: dict[int, dict[str, Any]] = {}
    for seq, topic in topic_map.items():
        finalized[seq] = {
            "seq": seq,
            "options": [
                topic["options"][option_type]
                for option_type in EXPECTED_OPTION_TYPES
                if option_type in topic["options"]
            ],
        }
    return finalized, len(fragment_paths)


def validate(topics: list[dict[str, Any]], fragments: dict[int, dict[str, Any]]) -> list[str]:
    errors: list[str] = []
    source_seqs = {item["seq"] for item in topics}
    fragment_seqs = set(fragments.keys())
    missing = sorted(source_seqs - fragment_seqs)
    extra = sorted(fragment_seqs - source_seqs)
    if missing:
        errors.append(f"缺少主题分片: {missing}")
    if extra:
        errors.append(f"存在多余主题分片: {extra}")

    seen_course_names: dict[str, str] = {}
    teacher_usage: defaultdict[str, list[str]] = defaultdict(list)
    for source in topics:
        seq = source["seq"]
        fragment = fragments.get(seq)
        if not fragment:
            continue
        options = fragment.get("options", [])
        if len(options) != len(EXPECTED_OPTION_TYPES):
            errors.append(f"seq={seq} 课程方案数量不是{len(EXPECTED_OPTION_TYPES)}个")
            continue
        option_types = sorted(normalize_text(item.get("option_type")) for item in options)
        if option_types != sorted(EXPECTED_OPTION_TYPES):
            errors.append(
                f"seq={seq} 方案类型必须包含{'/'.join(EXPECTED_OPTION_TYPES)}，当前为{option_types}"
            )
        for option in options:
            option_type = normalize_text(option.get("option_type"))
            for field in REQUIRED_OPTION_FIELDS:
                if not normalize_text(option.get(field)):
                    errors.append(f"seq={seq} option={option_type} 缺少字段 {field}")
            course_name = normalize_text(option.get("course_name"))
            if course_name:
                prev_seq = seen_course_names.get(course_name)
                if prev_seq is not None:
                    errors.append(f"课程名称重复：{prev_seq} 与 {seq}{option_type} -> {course_name}")
                else:
                    seen_course_names[course_name] = f"{seq}{option_type}"
            teacher_name = normalize_text(option.get("teacher_name"))
            if teacher_name:
                teacher_usage[teacher_name].append(f"{seq}{option_type}")
        if len({normalize_text(item.get('course_name')) for item in options}) != len(
            EXPECTED_OPTION_TYPES
        ):
            errors.append(f"seq={seq} 的四个方案中存在课程名称重复")

    overloaded_teachers = {name: seqs for name, seqs in teacher_usage.items() if len(seqs) > 4}
    if overloaded_teachers:
        errors.append(f"单名讲师覆盖超过4门课: {overloaded_teachers}")
    return errors


def markdown_escape(text: str) -> str:
    return text.replace("\n", "\n\n")


def append_section(text: str, title: str, items: list[str]) -> str:
    base = normalize_text(text)
    section_lines = [title]
    for index, item in enumerate(items, start=1):
        section_lines.append(f"{index}. {item}")
    addition = "\n".join(section_lines)
    if not base:
        return addition
    if addition in base:
        return base
    return f"{base}\n{addition}"


def enhance_outline(source: dict[str, Any], option: dict[str, Any]) -> str:
    outline = normalize_text(option["course_outline"])
    requirement = PROJECT_REQUIREMENTS[source["project"]]
    if not requirement["requires_hands_on"]:
        return outline

    title = OPTION_HANDS_ON_SECTION_TITLE[normalize_text(option["option_type"])]
    focus = requirement["practice_focus"]
    output_focus = OPTION_DELIVERY_FOCUS[normalize_text(option["option_type"])]
    items = [
        f"课程采用现场授课与上机实操演练相结合的方式开展，讲师围绕{focus}进行分步演示并组织学员完成{output_focus}。",
        "供应商需在我司培训场地布设或提供可供电脑上机实操的环境，保障学员按课堂节奏完成账号、工具、数据或实验资源的现场使用。",
        "课堂结束前完成实操结果复盘、问题答疑与成果验证，确保学员能够将训练内容迁移到后续项目实施或岗位实践中。",
    ]
    return append_section(outline, title, items)


def enhance_intro(source: dict[str, Any], option: dict[str, Any]) -> str:
    intro = normalize_text(option["course_intro"])
    requirement = PROJECT_REQUIREMENTS[source["project"]]
    prefix = (
        f"结合本培训项目面向{requirement['training_audience']}的线下培训定位，"
        f"本课程同步服务“{requirement['training_objective']}”这一项目级目标。"
    )
    if requirement["requires_hands_on"]:
        delivery = (
            "课程采用现场授课与上机实操演练相结合的方式开展，"
            "供应商需在我司培训场地布设或提供可供电脑上机实操的环境，"
            f"并由讲师现场指导学员围绕{requirement['practice_focus']}完成{OPTION_DELIVERY_FOCUS[normalize_text(option['option_type'])]}。"
        )
    else:
        delivery = (
            "课程按线下授课模式组织，可结合案例研讨、情景演练或操作示范推进学习，"
            "帮助学员把项目级培训目标落实到具体业务场景、岗位方法与协同动作中。"
        )
    if prefix in intro and delivery in intro:
        return intro
    return f"{intro}{prefix}{delivery}"


def enhance_teacher_bio(source: dict[str, Any], option: dict[str, Any]) -> str:
    bio = normalize_text(option["teacher_bio"])
    requirement = PROJECT_REQUIREMENTS[source["project"]]
    if not requirement["requires_hands_on"]:
        return bio

    teacher_standard = requirement["teacher_standard"]
    for prefix in ("讲师应具备", "讲师需具备", "讲师需为当前"):
        if teacher_standard.startswith(prefix):
            teacher_standard = teacher_standard[len(prefix) :]
            break
    teacher_standard = teacher_standard.rstrip("。")
    pedigree_text = f"该候选讲师具备{teacher_standard}。"
    delivery_text = requirement["teacher_delivery_suffix"]
    enhanced = bio
    if pedigree_text not in enhanced:
        enhanced = f"{pedigree_text}{enhanced}"
    if delivery_text and delivery_text not in enhanced:
        enhanced = f"{enhanced}{delivery_text}"
    return enhanced


def extract_outline_headings(outline: str) -> list[str]:
    text = normalize_text(outline).replace(" / ", "\n")
    headings = re.findall(
        r"(?:^|\n)(模块[一二三四五六七八九十0-9]+[:：][^\n]+|Day\s*\d+[:：][^\n]+|[一二三四五六七八九十]+[、:：][^\n]+)",
        text,
    )
    cleaned: list[str] = []
    for item in headings:
        value = re.sub(r"^(模块[一二三四五六七八九十0-9]+[:：]|Day\s*\d+[:：]|[一二三四五六七八九十]+[、:：])", "", item).strip()
        if value and value not in cleaned:
            cleaned.append(value)
    return cleaned[:3]


COURSE_OUTPUTS = {
    "AI应用开发": {
        "A": ["AI应用建设方法框架", "系统架构蓝图草案", "质量与安全检查清单"],
        "B": ["上机联调记录", "实操问题清单与修正建议", "应用实施步骤说明"],
        "C": ["专题优化方案", "专项架构治理清单", "性能或安全改进建议"],
        "D": ["分组共创方案", "实施路线图", "答辩优化纪要"],
    },
    "AI应用解决方案": {
        "A": ["场景识别清单", "能力地图", "机会判断表"],
        "B": ["需求转化草案", "场景卡片", "推进建议清单"],
        "C": ["专题能力图谱", "业务建模方案", "岗位进阶行动表"],
        "D": ["机会地图", "场景优先级清单", "共创行动计划"],
    },
    "AI应用运营": {
        "A": ["数据分析框架卡", "指标拆解表", "问题诊断路径图"],
        "B": ["看板设计草案", "异常诊断结论", "优化动作清单"],
        "C": ["专题复盘报告", "增长机会清单", "指标归因模型"],
        "D": ["专题诊断报告", "优化方案草案", "汇报材料提纲"],
    },
    "AI模型算法工程化": {
        "A": ["模型选型评估框架", "场景适配判断表", "风险清单"],
        "B": ["试点评测方案", "上机实验记录", "落地建议书"],
        "C": ["专题技术评估报告", "架构或性能优化清单", "实验结论摘要"],
        "D": ["技术路线建议书", "评审模板", "分组答辩成果"],
    },
    "AI数据能力提升": {
        "A": ["数据架构方法框架", "建模评审清单", "治理控制清单"],
        "B": ["数据工程实施记录", "服务设计草案", "质量问题清单"],
        "C": ["平台升级专题方案", "非结构化治理策略", "服务优化建议"],
        "D": ["实施方案文档", "服务发布清单", "答辩优化纪要"],
    },
    "AI能力全员赋能": {
        "A": ["AI工具认知图", "岗位应用清单", "合规提示卡"],
        "B": ["提示词模板", "轻量Agent配置记录", "个人提效清单"],
        "C": ["岗位场景应用方案", "效果评估清单", "安全使用规范"],
        "D": ["部门Agent需求卡", "原型演示结果", "试点推广计划"],
    },
}

SEMINAR_TEMPLATES = {
    "A": ["小组讨论", "案例分析", "互动演练"],
    "B": ["案例分析", "实战演练", "项目实战"],
    "C": ["小组讨论", "案例分析", "互动演练"],
    "D": ["小组讨论", "项目实战", "角色扮演"],
}

HIGHLIGHT_TEMPLATES = {
    "A": ["紧扣招标大纲中的核心知识点，形成结构化方法框架。", "兼顾金融科技业务场景与岗位职责，便于课后直接迁移应用。"] ,
    "B": ["案例与实战占比较高，强调操作步骤、结果验证与问题复盘。", "课堂产出物明确，便于培训后快速转化为项目执行动作。"] ,
    "C": ["专题切入更深，突出关键机制、专项能力与进阶优化方法。", "兼顾技术深度与金融科技场景适配，适合骨干人员能力提升。"] ,
    "D": ["采用共创式组织方式，强调分工协作、成果沉淀与现场答辩。", "课堂交付物导向明确，可直接服务后续立项、评审或项目推进。"] ,
}

APPLICABILITY_TEMPLATES = {
    "AI应用开发": "内容紧贴金融科技团队在低代码开发、模型集成、系统对接、安全验证等场景下的实际需求，适合用于提升机构级AI应用研发交付与质量保障能力。",
    "AI应用解决方案": "内容契合金融科技条线在场景识别、需求分析、业务建模和跨部门协同中的工作特点，能够缩短从业务想法到AI方案落地的转化路径。",
    "AI应用运营": "内容适配金融科技运营团队在智能运维、性能监控、指标分析和风险控制中的真实工作场景，可直接支撑运营决策与价值验证。",
    "AI模型算法工程化": "内容适配金融科技团队在模型选型、训练调优、部署上线和生命周期管理中的技术要求，有助于提升机构自有模型工程能力与落地效率。",
    "AI数据能力提升": "内容紧贴金融科技机构在数据治理、数据工程、非结构化处理和安全管控方面的建设需求，可支撑数据质量提升与数据资产价值释放。",
    "AI能力全员赋能": "内容适配金融科技机构的全员AI认知普及与场景推广需要，有助于建立统一的AI使用语言、合规意识和协同提效方式。",
}


def build_course_outputs(source: dict[str, Any], option_type: str) -> list[str]:
    return COURSE_OUTPUTS[source["project"]][option_type]


def build_seminar_arrangement(source: dict[str, Any], option: dict[str, Any]) -> list[str]:
    option_type = normalize_text(option["option_type"])
    headings = extract_outline_headings(option["course_outline"])
    topic_ref = "、".join(headings[:2]) if headings else source["project_practice_focus"]
    practice_focus = source["project_practice_focus"]
    arrangements = []
    for item in SEMINAR_TEMPLATES[option_type]:
        if item == "小组讨论":
            arrangements.append(f"小组讨论：围绕{source['topic']}对应的关键任务，分组讨论{topic_ref}在实际金融科技场景中的落地方式。")
        elif item == "案例分析":
            arrangements.append(f"案例分析：结合金融科技典型业务案例，拆解{topic_ref or practice_focus}相关问题、方法和常见风险。")
        elif item == "互动演练":
            arrangements.append(f"互动演练：通过现场问答、步骤推演或工具演示，帮助学员理解{practice_focus}中的关键控制点。")
        elif item == "实战演练":
            arrangements.append(f"实战演练：围绕课程核心模块组织现场操作或过程演练，验证学员对{practice_focus}方法的掌握程度。")
        elif item == "项目实战":
            arrangements.append(f"项目实战：以拟真项目任务推进课堂活动，要求学员形成可复用的方案草稿、记录或结论输出。")
        elif item == "角色扮演":
            arrangements.append("角色扮演：模拟业务、产品、研发、运营等多角色协作过程，检验方案沟通与任务分工能力。")
    return arrangements[:3]


def build_course_highlights(source: dict[str, Any], option: dict[str, Any]) -> list[str]:
    option_type = normalize_text(option["option_type"])
    return HIGHLIGHT_TEMPLATES[option_type][:2]


def build_applicability_value(source: dict[str, Any], option: dict[str, Any]) -> str:
    base = APPLICABILITY_TEMPLATES[source["project"]]
    option_type = normalize_text(option["option_type"])
    if option_type == "D":
        suffix = "同时，课程通过共创与成果输出方式，更适合用于机构内部跨团队共识对齐和项目启动前的能力准备。"
    elif option_type == "C":
        suffix = "对于承担专项攻关、体系升级或深度优化任务的人员，这类专题化内容更有助于形成可持续的专业能力积累。"
    elif option_type == "B":
        suffix = "通过案例和演练驱动的方式，课程成果更容易映射到日常项目执行、问题处置和效果验证工作。"
    else:
        suffix = "通过方法框架的统一讲解，课程有助于不同角色建立一致的理解口径和执行标准。"
    return f"{base}{suffix}"


def compose_course_intro(
    course_description: str,
    course_outputs: list[str],
    seminar_arrangement: list[str],
    course_highlights: list[str],
    applicability_value: str,
) -> str:
    sections = [
        f"课程介绍：{course_description}",
        "课程产出：" + "".join(f"{idx}. {item}" for idx, item in enumerate(course_outputs, start=1)),
        "课程研讨安排：" + "".join(f"{idx}. {item}" for idx, item in enumerate(seminar_arrangement, start=1)),
        "课程亮点：" + "".join(f"{idx}. {item}" for idx, item in enumerate(course_highlights, start=1)),
        f"适用性和价值度：{applicability_value}",
    ]
    return "\n".join(sections)


def adapt_option(source: dict[str, Any], option: dict[str, Any]) -> dict[str, Any]:
    course_description = enhance_intro(source, option)
    course_outputs = build_course_outputs(source, normalize_text(option["option_type"]))
    seminar_arrangement = build_seminar_arrangement(source, option)
    course_highlights = build_course_highlights(source, option)
    applicability_value = build_applicability_value(source, option)
    adapted = {
        "option_type": normalize_text(option["option_type"]),
        "course_name": normalize_text(option["course_name"]),
        "course_outline": enhance_outline(source, option),
        "course_description": course_description,
        "course_outputs": course_outputs,
        "seminar_arrangement": seminar_arrangement,
        "course_highlights": course_highlights,
        "applicability_value": applicability_value,
        "course_intro": compose_course_intro(
            course_description,
            course_outputs,
            seminar_arrangement,
            course_highlights,
            applicability_value,
        ),
        "teacher_name": normalize_text(option["teacher_name"]),
        "teacher_bio": enhance_teacher_bio(source, option),
    }
    return adapted


def build_candidate_records(
    source_topics: list[dict[str, Any]],
    fragment_map: dict[int, dict[str, Any]],
) -> tuple[list[dict[str, Any]], dict[str, list[dict[str, Any]]]]:
    source_by_seq = {item["seq"]: item for item in source_topics}
    candidate_records: list[dict[str, Any]] = []
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for seq in sorted(source_by_seq):
        source = source_by_seq[seq]
        fragment = fragment_map[seq]
        ordered_options = sorted(
            fragment["options"],
            key=lambda item: EXPECTED_OPTION_TYPES.index(normalize_text(item["option_type"])),
        )
        adapted_options = [adapt_option(source, option) for option in ordered_options]
        grouped[source["project"]].append({"source": source, "options": adapted_options})
        for option in adapted_options:
            candidate_records.append(
                {
                    "seq": seq,
                    "project": source["project"],
                    "topic": source["topic"],
                    "option_type": option["option_type"],
                    "course_name": option["course_name"],
                    "course_outline": option["course_outline"],
                    "course_description": option["course_description"],
                    "course_outputs": " | ".join(option["course_outputs"]),
                    "seminar_arrangement": " | ".join(option["seminar_arrangement"]),
                    "course_highlights": " | ".join(option["course_highlights"]),
                    "applicability_value": option["applicability_value"],
                    "course_intro": option["course_intro"],
                    "teacher_name": option["teacher_name"],
                    "teacher_bio": option["teacher_bio"],
                }
            )
    return candidate_records, grouped


def render_project_requirement_block(project: str) -> list[str]:
    requirement = PROJECT_REQUIREMENTS[project]
    lines = [
        "### 项目实施要求",
        "",
        f"- **培训方式**：拟通过{requirement['delivery_mode']}开展。",
        f"- **培训安排**：{requirement['session_schedule']}，{requirement['participant_range']}。",
        f"- **培训对象**：{requirement['training_audience']}",
        f"- **培训目标**：{requirement['training_objective']}",
        f"- **实施要求**：{requirement['delivery_requirement']}",
        f"- **师资要求**：{requirement['teacher_standard']}",
        "",
    ]
    return lines


def render_markdown(grouped: dict[str, list[dict[str, Any]]]) -> str:
    lines: list[str] = []
    lines.append("# AI培训课程设计与授课研讨安排")
    lines.append("")
    lines.append("> 说明：本稿按课程设计与授课研讨样式整理，每个课程主题提供四门候选课程，分别为“方法论版（A）”“实战版（B）”“专项深化版（C）”“方案共创版（D）”。授课教师为候选讲师方案，可在后续商务排期阶段替换为最终授课名单。")
    lines.append("")
    for project, items in grouped.items():
        lines.append(f"## {project}")
        lines.append("")
        lines.extend(render_project_requirement_block(project))
        for entry in items:
            source = entry["source"]
            lines.append(f"### 主题{source['seq']}：{source['topic']}")
            lines.append("")
            lines.append(f"- **培训对象**：{source['audience']}")
            lines.append(f"- **培训目标**：{source['objective']}")
            lines.append(f"- **项目级培训对象**：{source['project_training_audience']}")
            lines.append(f"- **项目级培训目标**：{source['project_training_objective']}")
            lines.append(f"- **参考时长**：{source['duration_normalized'] or source['duration_raw']}")
            lines.append("")
            for option in entry["options"]:
                option_title = OPTION_TITLES.get(option["option_type"], "候选版")
                lines.append(f"#### 备选课{option['option_type']}（{option_title}）：{option['course_name']}")
                lines.append("")
                lines.append(f"- **课程大纲**：{markdown_escape(normalize_text(option['course_outline']))}")
                lines.append(f"- **课程介绍**：{markdown_escape(normalize_text(option['course_description']))}")
                lines.append(f"- **课程产出**：{markdown_escape('；'.join(option['course_outputs']))}")
                lines.append(f"- **课程研讨安排**：{markdown_escape('；'.join(option['seminar_arrangement']))}")
                lines.append(f"- **授课教师**：{normalize_text(option['teacher_name'])}")
                lines.append(f"- **课程亮点**：{markdown_escape('；'.join(option['course_highlights']))}")
                lines.append(f"- **适用性和价值度**：{markdown_escape(normalize_text(option['applicability_value']))}")
                lines.append(f"- **授课教师简介**：{markdown_escape(normalize_text(option['teacher_bio']))}")
                lines.append("")
    return "\n".join(lines).strip() + "\n"


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    if not rows:
        return
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    topics = extract_topics()
    (OUTPUT_DIR / "normalized_topics.json").write_text(
        json.dumps(topics, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    fragments, fragment_file_count = load_fragments()
    errors = validate(topics, fragments)
    if errors:
        raise SystemExit("校验失败：\n- " + "\n- ".join(errors))

    candidate_records, grouped = build_candidate_records(topics, fragments)
    (OUTPUT_DIR / "candidate_courses.json").write_text(
        json.dumps(candidate_records, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    write_csv(OUTPUT_DIR / "candidate_courses.csv", candidate_records)
    (OUTPUT_DIR / "ai_training_course_draft.md").write_text(
        render_markdown(grouped),
        encoding="utf-8",
    )

    summary = {
        "topic_count": len(topics),
        "course_count": len(candidate_records),
        "project_count": len({topic["project"] for topic in topics}),
        "fragment_file_count": fragment_file_count,
        "fragment_topic_count": len(fragments),
    }
    (OUTPUT_DIR / "summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
