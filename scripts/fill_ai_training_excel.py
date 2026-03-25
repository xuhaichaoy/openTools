from __future__ import annotations

import argparse
import json
import shutil
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment


ROOT = Path("/Users/haichao/Desktop/work/51ToolBox")
DEFAULT_SOURCE = Path("/Users/haichao/Downloads/AI培训课程需求.xlsx")
DEFAULT_TOPICS = ROOT / "deliverables" / "ai_training_course_draft" / "normalized_topics.json"
DEFAULT_COURSES = ROOT / "deliverables" / "ai_training_course_draft" / "candidate_courses.json"
OPTION_ORDER = ("A", "B", "C", "D")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="回填 AI 培训课程候选方案到原始 Excel")
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--topics", type=Path, default=DEFAULT_TOPICS)
    parser.add_argument("--courses", type=Path, default=DEFAULT_COURSES)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--backup", type=Path, default=None)
    return parser.parse_args()


def load_json(path: Path):
    return json.loads(path.read_text(encoding="utf-8"))


def copy_row_format(ws, source_row: int, target_row: int, max_col: int) -> None:
    if ws.row_dimensions[source_row].height is not None:
        ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    ws.row_dimensions[target_row].hidden = ws.row_dimensions[source_row].hidden
    ws.row_dimensions[target_row].outlineLevel = ws.row_dimensions[source_row].outlineLevel

    for col_idx in range(1, max_col + 1):
        source_cell = ws.cell(source_row, col_idx)
        target_cell = ws.cell(target_row, col_idx)
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format
        if source_cell.font:
            target_cell.font = copy(source_cell.font)
        if source_cell.fill:
            target_cell.fill = copy(source_cell.fill)
        if source_cell.border:
            target_cell.border = copy(source_cell.border)
        if source_cell.alignment:
            target_cell.alignment = copy(source_cell.alignment)
        if source_cell.protection:
            target_cell.protection = copy(source_cell.protection)


def clear_values(ws, start_row: int, end_row: int, max_col: int) -> None:
    for row in range(start_row, end_row + 1):
        for col in range(1, max_col + 1):
            ws.cell(row, col).value = None


def ensure_wrapped(ws, row: int, columns: range) -> None:
    for col in columns:
        cell = ws.cell(row, col)
        current = cell.alignment or Alignment()
        cell.alignment = copy(
            Alignment(
                horizontal=current.horizontal or "left",
                vertical=current.vertical or "top",
                text_rotation=current.text_rotation,
                wrap_text=True,
                shrink_to_fit=current.shrink_to_fit,
                indent=current.indent,
            )
        )


def group_topic_ranges(topics: list[dict]) -> list[tuple[str, int, int]]:
    groups: list[tuple[str, int, int]] = []
    start_idx = 0
    while start_idx < len(topics):
        project = topics[start_idx]["project"]
        end_idx = start_idx
        while end_idx + 1 < len(topics) and topics[end_idx + 1]["project"] == project:
            end_idx += 1
        groups.append((project, start_idx, end_idx))
        start_idx = end_idx + 1
    return groups


def main() -> None:
    args = parse_args()
    topics: list[dict] = load_json(args.topics)
    course_records: list[dict] = load_json(args.courses)

    course_map: dict[int, dict[str, dict]] = {}
    for record in course_records:
        course_map.setdefault(int(record["seq"]), {})[record["option_type"]] = record
    option_types = [option for option in OPTION_ORDER if any(option in item for item in course_map.values())]
    rows_per_topic = len(option_types)
    if rows_per_topic == 0:
        raise SystemExit("未找到可回填课程方案")

    workbook = load_workbook(args.source)
    worksheet = workbook[workbook.sheetnames[0]]
    original_max_row = worksheet.max_row
    max_col = worksheet.max_column

    for merged in list(worksheet.merged_cells.ranges):
        worksheet.unmerge_cells(str(merged))

    for original_row in range(original_max_row, 2, -1):
        worksheet.insert_rows(original_row + 1, 1)
        copy_row_format(worksheet, original_row, original_row + 1, max_col)

    extra_rows_needed = len(topics) * (rows_per_topic - 2)
    if extra_rows_needed > 0:
        for _ in range(extra_rows_needed):
            insert_at = worksheet.max_row + 1
            worksheet.insert_rows(insert_at, 1)
            copy_row_format(worksheet, worksheet.max_row - 1, insert_at, max_col)

    final_max_row = 2 + len(topics) * rows_per_topic
    clear_values(worksheet, 3, final_max_row, max_col)

    for topic_index, topic in enumerate(topics):
        base_row = 3 + topic_index * rows_per_topic
        option_records = course_map[int(topic["seq"])]

        worksheet.cell(base_row, 1).value = topic["seq"]
        worksheet.cell(base_row, 4).value = topic["topic"]
        worksheet.cell(base_row, 5).value = topic["audience"]
        worksheet.cell(base_row, 6).value = topic["objective"]
        worksheet.cell(base_row, 7).value = topic["source_outline"]
        worksheet.cell(base_row, 8).value = topic["duration_raw"]
        worksheet.cell(base_row, 9).value = topic.get("teacher_requirement", "")

        for offset, option_type in enumerate(option_types):
            option_row = base_row + offset
            option = option_records[option_type]
            worksheet.cell(option_row, 10).value = option["course_name"]
            worksheet.cell(option_row, 11).value = option["course_outline"]
            worksheet.cell(option_row, 12).value = option["course_intro"]
            worksheet.cell(option_row, 13).value = option["teacher_name"]
            worksheet.cell(option_row, 14).value = option["teacher_bio"]
            worksheet.cell(option_row, 15).value = f"{topic['seq']}{option_type}"
            ensure_wrapped(worksheet, option_row, range(10, 15))

        ensure_wrapped(worksheet, base_row, range(4, 10))

        for column in [1, 4, 5, 6, 7, 8, 9]:
            worksheet.merge_cells(
                start_row=base_row,
                start_column=column,
                end_row=base_row + rows_per_topic - 1,
                end_column=column,
            )

    for _, start_idx, end_idx in group_topic_ranges(topics):
        start_row = 3 + start_idx * rows_per_topic
        end_row = start_row + (end_idx - start_idx + 1) * rows_per_topic - 1
        worksheet.cell(start_row, 2).value = topics[start_idx]["project"]
        worksheet.cell(start_row, 3).value = topics[start_idx]["project_period"]
        worksheet.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)
        worksheet.merge_cells(start_row=start_row, start_column=3, end_row=end_row, end_column=3)

    worksheet.merge_cells("A1:I1")
    worksheet.merge_cells("J1:O1")

    if args.backup:
        args.backup.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(args.source, args.backup)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(args.output)
    print(args.output)


if __name__ == "__main__":
    main()
