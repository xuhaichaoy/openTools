from __future__ import annotations

import argparse
import re
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


TOPIC_PATTERN = re.compile(r"^# 主题(\d+)：(.+)$")
COURSE_PATTERN = re.compile(r"^#{2,6}\s*课程([一二三四五六七八九十0-9]+)：(.+)$")
SCALAR_PATTERN = re.compile(r"^- \*\*(.+?)\*\*：\s*(.*)$")
LIST_ITEM_PATTERN = re.compile(r"^\s+\d+\.\s*(.+)$")

TOPIC_SCALAR_MAP = {
    "项目类别": "project_category",
    "授课对象": "audience",
    "培训目标": "objective",
}

COURSE_SCALAR_MAP = {
    "课程主题": "course_theme",
    "课程介绍": "course_intro",
    "适用性和价值度": "applicability_value",
}

COURSE_LIST_MAP = {
    "课程产出": "deliverables",
    "课程研讨安排": "discussion_plan",
    "课程亮点": "highlights",
    "覆盖确认": "coverage_confirmation",
}


@dataclass
class CourseRecord:
    course_order_label: str = ""
    course_order: int = 0
    course_heading: str = ""
    course_theme: str = ""
    course_intro: str = ""
    deliverables: list[str] = field(default_factory=list)
    discussion_plan: list[str] = field(default_factory=list)
    highlights: list[str] = field(default_factory=list)
    applicability_value: str = ""
    coverage_confirmation: list[str] = field(default_factory=list)


@dataclass
class TopicRecord:
    topic_seq: int
    topic_name: str
    project_category: str = ""
    audience: str = ""
    objective: str = ""
    courses: list[CourseRecord] = field(default_factory=list)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="将五门课程 Markdown 成稿导出为 Excel")
    parser.add_argument("--input", type=Path, required=True, help="输入 Markdown 文件路径")
    parser.add_argument("--output", type=Path, required=True, help="输出 Excel 文件路径")
    parser.add_argument(
        "--mode",
        choices=("raw", "structured"),
        default="raw",
        help="raw 为完整 Markdown 原文导出，structured 为按主题与课程结构化拆分导出",
    )
    return parser.parse_args()


def clean_text(text: str) -> str:
    return text.strip()


def append_to_last_item(items: list[str], line: str) -> None:
    if not items:
        items.append(clean_text(line))
        return
    items[-1] = f"{items[-1]} {clean_text(line)}".strip()


def append_to_scalar(existing: str, line: str) -> str:
    if not existing:
        return clean_text(line)
    return f"{existing} {clean_text(line)}".strip()


def parse_markdown(path: Path) -> list[TopicRecord]:
    topics: list[TopicRecord] = []
    current_topic: TopicRecord | None = None
    current_course: CourseRecord | None = None
    current_section: str | None = None

    lines = path.read_text(encoding="utf-8").splitlines()
    for raw_line in lines:
        line = raw_line.rstrip()

        if not line.strip() or line.startswith("> "):
            continue

        topic_match = TOPIC_PATTERN.match(line)
        if topic_match:
            if current_course and current_topic:
                current_topic.courses.append(current_course)
                current_course = None
            if current_topic:
                topics.append(current_topic)
            current_topic = TopicRecord(
                topic_seq=int(topic_match.group(1)),
                topic_name=clean_text(topic_match.group(2)),
            )
            current_section = None
            continue

        course_match = COURSE_PATTERN.match(line)
        if course_match:
            if not current_topic:
                raise ValueError(f"发现课程段落但尚未进入主题：{line}")
            if current_course:
                current_topic.courses.append(current_course)
            current_course = CourseRecord(
                course_order_label=course_match.group(1),
                course_order=len(current_topic.courses) + 1,
                course_heading=clean_text(course_match.group(2)),
            )
            current_section = None
            continue

        scalar_match = SCALAR_PATTERN.match(line)
        if scalar_match:
            key = clean_text(scalar_match.group(1))
            value = clean_text(scalar_match.group(2))
            if key in TOPIC_SCALAR_MAP:
                if not current_topic:
                    raise ValueError(f"发现主题属性但尚未进入主题：{line}")
                setattr(current_topic, TOPIC_SCALAR_MAP[key], value)
                current_section = TOPIC_SCALAR_MAP[key]
                continue
            if key in COURSE_SCALAR_MAP:
                if not current_course:
                    raise ValueError(f"发现课程属性但尚未进入课程：{line}")
                setattr(current_course, COURSE_SCALAR_MAP[key], value)
                current_section = COURSE_SCALAR_MAP[key]
                continue
            if key in COURSE_LIST_MAP:
                if not current_course:
                    raise ValueError(f"发现课程列表段落但尚未进入课程：{line}")
                current_section = COURSE_LIST_MAP[key]
                continue

        item_match = LIST_ITEM_PATTERN.match(line)
        if item_match and current_course and current_section in COURSE_LIST_MAP.values():
            getattr(current_course, current_section).append(clean_text(item_match.group(1)))
            continue

        if line.startswith("# ") or line.startswith("## "):
            continue

        if current_course and current_section in COURSE_LIST_MAP.values():
            append_to_last_item(getattr(current_course, current_section), line)
            continue

        if current_course and current_section in COURSE_SCALAR_MAP.values():
            current_value = getattr(current_course, current_section)
            setattr(current_course, current_section, append_to_scalar(current_value, line))
            continue

        if current_topic and current_section in TOPIC_SCALAR_MAP.values():
            current_value = getattr(current_topic, current_section)
            setattr(current_topic, current_section, append_to_scalar(current_value, line))
            continue

    if current_course and current_topic:
        current_topic.courses.append(current_course)
    if current_topic:
        topics.append(current_topic)

    return topics


def join_lines(items: list[str]) -> str:
    return "\n".join(f"{index}. {item}" for index, item in enumerate(items, start=1))


def style_header_row(worksheet, row: int, column_count: int) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )
    for column in range(1, column_count + 1):
        cell = worksheet.cell(row, column)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border


def apply_body_style(worksheet, start_row: int, end_row: int, column_count: int) -> None:
    border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    for row in range(start_row, end_row + 1):
        for column in range(1, column_count + 1):
            cell = worksheet.cell(row, column)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border


def build_topic_overview_sheet(workbook: Workbook, topics: list[TopicRecord]) -> None:
    worksheet = workbook.active
    worksheet.title = "主题概览"

    headers = ["主题序号", "主题名称", "项目类别", "授课对象", "培训目标", "课程数量"]
    worksheet.append(headers)
    style_header_row(worksheet, 1, len(headers))

    for topic in topics:
        worksheet.append(
            [
                topic.topic_seq,
                topic.topic_name,
                topic.project_category,
                topic.audience,
                topic.objective,
                len(topic.courses),
            ]
        )

    apply_body_style(worksheet, 2, worksheet.max_row, len(headers))
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    widths = [10, 52, 18, 28, 60, 12]
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width


def build_course_detail_sheet(workbook: Workbook, topics: list[TopicRecord]) -> None:
    worksheet = workbook.create_sheet("课程明细")
    headers = [
        "主题序号",
        "主题名称",
        "项目类别",
        "授课对象",
        "培训目标",
        "课程序号",
        "课程标题",
        "课程介绍",
        "课程产出",
        "课程研讨安排",
        "课程亮点",
        "适用性和价值度",
        "覆盖确认",
    ]
    worksheet.append(headers)
    style_header_row(worksheet, 1, len(headers))

    for topic in topics:
        for course in topic.courses:
            worksheet.append(
                [
                    topic.topic_seq,
                    topic.topic_name,
                    topic.project_category,
                    topic.audience,
                    topic.objective,
                    course.course_order,
                    course.course_theme or course.course_heading,
                    course.course_intro,
                    join_lines(course.deliverables),
                    join_lines(course.discussion_plan),
                    join_lines(course.highlights),
                    course.applicability_value,
                    join_lines(course.coverage_confirmation),
                ]
            )

    apply_body_style(worksheet, 2, worksheet.max_row, len(headers))
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    widths = [10, 52, 18, 28, 60, 10, 34, 72, 34, 42, 42, 42, 42]
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width


def build_export_info_sheet(workbook: Workbook, input_path: Path, topics: list[TopicRecord]) -> None:
    worksheet = workbook.create_sheet("导出说明")
    rows = [
        ("来源文件", str(input_path)),
        ("主题数量", len(topics)),
        ("课程数量", sum(len(topic.courses) for topic in topics)),
        ("导出说明", "课程明细按一门课程一行展开，多项内容以单元格内换行呈现，便于二次编辑和筛选。"),
    ]
    for row in rows:
        worksheet.append(row)

    for column in range(1, 3):
        worksheet.column_dimensions[get_column_letter(column)].width = 18 if column == 1 else 120
    apply_body_style(worksheet, 1, worksheet.max_row, 2)
    for row in range(1, worksheet.max_row + 1):
        worksheet.cell(row, 1).font = Font(bold=True)
    worksheet.freeze_panes = "A2"


def build_raw_markdown_sheet(workbook: Workbook, input_path: Path) -> None:
    worksheet = workbook.active
    worksheet.title = "完整MD原文"
    headers = ["行号", "Markdown内容"]
    worksheet.append(headers)
    style_header_row(worksheet, 1, len(headers))

    lines = input_path.read_text(encoding="utf-8").splitlines()
    for index, line in enumerate(lines, start=1):
        worksheet.append([index, line])

    apply_body_style(worksheet, 2, worksheet.max_row, len(headers))
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.column_dimensions["A"].width = 10
    worksheet.column_dimensions["B"].width = 160


def build_raw_chunk_sheet(workbook: Workbook, input_path: Path) -> None:
    worksheet = workbook.create_sheet("原文分段")
    headers = ["段序号", "字符数", "Markdown原文分段"]
    worksheet.append(headers)
    style_header_row(worksheet, 1, len(headers))

    text = input_path.read_text(encoding="utf-8")
    chunk_size = 30000
    for index, start in enumerate(range(0, len(text), chunk_size), start=1):
        chunk = text[start : start + chunk_size]
        worksheet.append([index, len(chunk), chunk])

    apply_body_style(worksheet, 2, worksheet.max_row, len(headers))
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.column_dimensions["A"].width = 10
    worksheet.column_dimensions["B"].width = 12
    worksheet.column_dimensions["C"].width = 140


def build_raw_info_sheet(workbook: Workbook, input_path: Path) -> None:
    worksheet = workbook.create_sheet("导出说明")
    text = input_path.read_text(encoding="utf-8")
    rows = [
        ("来源文件", str(input_path)),
        ("导出模式", "完整 Markdown 原文导出"),
        ("总字符数", len(text)),
        ("总行数", len(text.splitlines())),
        ("说明", "完整原文按逐行方式写入工作表，另提供按 30000 字符切分的分段页，避免超出 Excel 单元格长度限制。"),
    ]
    for row in rows:
        worksheet.append(row)

    for column in range(1, 3):
        worksheet.column_dimensions[get_column_letter(column)].width = 18 if column == 1 else 120
    apply_body_style(worksheet, 1, worksheet.max_row, 2)
    for row in range(1, worksheet.max_row + 1):
        worksheet.cell(row, 1).font = Font(bold=True)
    worksheet.freeze_panes = "A2"


def export_structured_to_excel(input_path: Path, output_path: Path) -> None:
    topics = parse_markdown(input_path)
    if not topics:
        raise ValueError(f"未从文件中解析出主题：{input_path}")

    workbook = Workbook()
    build_topic_overview_sheet(workbook, topics)
    build_course_detail_sheet(workbook, topics)
    build_export_info_sheet(workbook, input_path, topics)
    workbook.save(output_path)


def export_raw_to_excel(input_path: Path, output_path: Path) -> None:
    workbook = Workbook()
    build_raw_markdown_sheet(workbook, input_path)
    build_raw_chunk_sheet(workbook, input_path)
    build_raw_info_sheet(workbook, input_path)
    workbook.save(output_path)


def main() -> None:
    args = parse_args()
    args.output.parent.mkdir(parents=True, exist_ok=True)
    if args.mode == "structured":
        export_structured_to_excel(args.input, args.output)
    else:
        export_raw_to_excel(args.input, args.output)
    print(args.output)


if __name__ == "__main__":
    main()
